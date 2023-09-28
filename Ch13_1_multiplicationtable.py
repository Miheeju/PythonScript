import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']

bold_font = Font(bold = True)

def multiplication_table(n):
    for i in range(1, n+1):
        # Set row index in bold
        sheet.cell(row= i + 1, column=1).font = bold_font
        sheet.cell(row= i+1, column=1).value = i
        # Set column headers in bold
        sheet.cell(row=1, column=i+1).font = bold_font
        sheet.cell(row=1, column=i+1).value = i
    # Now that headers are set, fill in center.
    # Need to specify i again otherwise some headers are NoneType and cannot be populated.
    for i in range(1, n + 1):
        for j in range(1, n+1):
            # row_value = sheet.cell(row=1, column=j + 1).value
            # column_value = sheet.cell(row=i + 1, column=1).value
            sheet.cell(row = i+1, column = j+1).value = i * j #row_value * column_value

multiplication_table(6)

wb.save('multiply_table.xlsx')
