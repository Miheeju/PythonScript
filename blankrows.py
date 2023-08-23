'''
At row N, insert M blank rows into spreadsheet
Read spreadsheet contents
Write to new spreadsheet with for loop to copy first N lines
Insert M blank rows
Then copy row from # N + M to max_row
'''

import openpyxl

# Read workbook
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active

# Open a new workbook
wb2 = openpyxl.Workbook()
sheet2 = wb2['Sheet'] #not wb['Sheet']!

def blank_row_inserter(n, m):
# Copy rows 1 to n
# j column range is same for all scenarios so use outermost indent
    for j in range(1, sheet.max_column):
        for i in range(1, n):
            sheet2.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value
# Fill in blank rows
        for i_blank in range(n, n+m):
            sheet2.cell(row=i_blank, column=j).value = None
# Copy remaining rows
        for i in range(n+m, sheet.max_row):
            sheet2.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value

    wb2.save('blank_rows.xlsx') # Not wb.save, use wb2.save! Keep inside function

blank_row_inserter(5, 2)